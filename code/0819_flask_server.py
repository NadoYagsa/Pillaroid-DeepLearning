#!/usr/bin/env python
# coding: utf-8

# In[ ]:


from flask import Flask, make_response, request, jsonify
import cv2
import os
import openpyxl

import sys
sys.path.append(os.path.dirname(os.path.abspath(os.getcwd())))

from yolov5 import detect

from keras.models import load_model
from keras.preprocessing import image
import numpy as np

app = Flask(__name__)
model = None

pill_filename = "datafile/pot_dataset_label.xlsx"


# 알약 인식 모델 로드
def load_pill_model():
    global model
    model = load_model('model_adam_ver2_epoch_37.h5')
    
def find_pill_number(pillLabel):
    # 엑셀 파일 읽기
    wb = openpyxl.load_workbook(pill_filename, data_only=True)

    sheet_list = wb.sheetnames

    w = wb[sheet_list[0]]

    serial_number = w.cell(row=pillLabel+2, column=1).value
    
    return serial_number
    

# 조회 이미지 전처리
def processing_image(pill_image, target):
    img = pill_image.resize(target)
    img_array = image.img_to_array(img)
    img_array /= 255.
    img_array = np.expand_dims(img_array, axis=0)

    return img_array


# HTTP POST방식으로 전송된 이미지를 저장
@app.route('/pill/predict', methods=['POST'])
def search_pill():
    data = {}
    # file 확인
    if request.files.get("image"):
        # PIL 형식으로 이미지 읽기
        img = request.files["image"].read()
        img = np.frombuffer(img, dtype=np.uint8)
        img = cv2.imdecode(img, cv2.IMREAD_COLOR)
        
        # 이미지에서 알약 영역을 자름
        img = detect.run_from_img_object(im0s=img, weights="../yolov5/runs/train/camera_web3/weights/best.pt", save_crop=True, nosave=True)

        # 크롭된 알약 영역이 없을 시(인식된 알약이 없을 시)
        if img == None:
            data["success"] = False
            return make_response(jsonify(data), 400)

        # 이미지를 전처리
        img = processing_image(img, target=(200, 200))

        # 이미지를 분류
        results = model.predict(img)
        data["predictions"] = []
        
        dict_results = {index : prob for index,prob in enumerate(results[0])}
        sorted_results = sorted(dict_results.items(), key = lambda item: item[1], reverse = True)

        # 예측 목록 저장
        for (label, prob) in sorted_results[0:5]:
            candidate = {"serial_number": find_pill_number(label), "probability": float(prob)}
            data["predictions"].append(candidate)

        data["success"] = True

        # Json 데이터 반환
        return jsonify(data)
    # 이미지가 없을 시
    else:
        data["success"] = False

        return make_response(jsonify(data), 404)


    
# 메인 쓰레드: 모델 로드 및 서버 실행
if __name__ == "__main__":
    print("* Loading Keras model and Flask starting server...")
    load_pill_model()
    app.run(host='0.0.0.0', port=5000)

